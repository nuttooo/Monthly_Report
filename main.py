import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

# ตั้งค่าพาธของโฟลเดอร์
data_folder = 'data/'

# ฟังก์ชันสำหรับให้ผู้ใช้เลือกโปรเจค
def select_project():
    projects = sorted([name for name in os.listdir(data_folder) if os.path.isdir(os.path.join(data_folder, name))])
    print("เลือกโปรเจค:")
    for i, project in enumerate(projects, 1):
        print(f"{i}. {project}")
    choice = int(input("กรุณาเลือกโปรเจค (ระบุหมายเลข): "))
    if 1 <= choice <= len(projects):
        return projects[choice - 1]
    else:
        print("กรุณาเลือกโปรเจคให้ถูกต้อง")
        return select_project()

# ฟังก์ชันสำหรับให้ผู้ใช้เลือกว่าจะใช้ข้อมูล PEA หรือ MEA
def select_pea_or_mea():
    print("เลือกหน่วยงาน:")
    print("1. PEA")
    print("2. MEA")
    choice = input("กรุณาเลือก (1 หรือ 2): ")
    if choice == '1':
        return "PEA"
    elif choice == '2':
        return "MEA"
    else:
        print("กรุณาเลือกให้ถูกต้อง (1 หรือ 2)")
        return select_pea_or_mea()

# กำหนดวันหยุดของ PEA
pea_holidays = {
    "2024-01-01", "2024-01-15", "2024-05-01", "2024-05-22", "2024-06-03", "2024-08-12", 
    "2024-10-13", "2024-10-23", "2024-12-05", "2024-12-10", "2024-12-31"
}

# กำหนดวันหยุดของ MEA
mea_holidays = {
    "2024-01-01", "2024-02-24", "2024-04-06", "2024-04-13", "2024-04-14", "2024-04-15",
    "2024-05-01", "2024-05-04", "2024-05-22", "2024-06-03", "2024-07-20", "2024-07-21",
    "2024-07-28", "2024-08-12", "2024-10-13", "2024-10-23", "2024-12-05", "2024-12-10",
    "2024-12-31"
}

# กำหนดสีสำหรับการไฮไลต์วันหยุดและวันเสาร์-อาทิตย์
pea_holiday_fill = PatternFill(start_color="7ACB4D", end_color="7ACB4D", fill_type="solid") 
pea_weekend_fill = PatternFill(start_color="7ACB4D", end_color="7ACB4D", fill_type="solid")  
mea_holiday_fill = PatternFill(start_color="FF905E", end_color="FF905E", fill_type="solid") 
mea_weekend_fill = PatternFill(start_color="757575", end_color="757575", fill_type="solid")  
mea_weekday_holiday_fill = PatternFill(start_color="CB501C", end_color="CB501C", fill_type="solid") 

# ฟังก์ชันคำนวณ To Home (kWh)
def calculate_to_home(inverter_yield, export):
    try:
        return inverter_yield - export
    except:
        return None

# เริ่มต้นโค้ดหลัก
selected_project = select_project()  # ให้ผู้ใช้เลือกโปรเจค
project_path = os.path.join(data_folder, selected_project)  # กำหนดพาธของโปรเจคที่เลือก

# กำหนดช่วงเวลา Onpeak และ Offpeak
onpeak_start = pd.to_datetime('09:00:00').time()
onpeak_end = pd.to_datetime('22:00:00').time()

selection = select_pea_or_mea()  # ให้ผู้ใช้เลือกหน่วยงาน

# เลือกวันหยุดและสีตามการเลือก
if selection == "PEA":
    holidays = pea_holidays
    holiday_fill = pea_holiday_fill
    weekend_fill = pea_weekend_fill
else:
    holidays = mea_holidays
    holiday_fill = mea_holiday_fill
    weekend_fill = mea_weekend_fill

output_file = f'output/{selected_project}_Monthly_Report.xlsx'  # ตั้งชื่อไฟล์ตามชื่อโปรเจค

# สร้างลิสต์สำหรับเก็บข้อมูลทุกวัน
all_days_data = []
summary_data = []

# ประมวลผลทุกไฟล์ในโฟลเดอร์ของโปรเจ็คนั้นๆ
for filename in sorted(os.listdir(project_path)):
    if filename.endswith('.xlsx'):
        # อ่านข้อมูลจากไฟล์
        file_path = os.path.join(project_path, filename)
        df = pd.read_excel(file_path, skiprows=1)

        # แปลงคอลัมน์ 'Statistical Period' ให้เป็น datetime
        df['Statistical Period'] = pd.to_datetime(df['Statistical Period'], errors='coerce')

        # คัดเลือกเฉพาะคอลัมน์เวลาที่เป็น A และค่า Inverter Yield กับ Export
        time_column = df['Statistical Period'].dt.time
        df_selected = df[[' Inverter Yield (kWh)', 'Export (kWh)']].copy()

        # คำนวณ To Home (kWh)
        df_selected['To Home (kWh)'] = df_selected.apply(
            lambda row: calculate_to_home(row[' Inverter Yield (kWh)'], row['Export (kWh)']), axis=1
        )

        # เพิ่มคอลัมน์เวลา
        df_selected.insert(0, 'Time', time_column)

        # ดึงวันที่จากชื่อไฟล์
        day = filename.split('_')[1].split('.')[0]

        # ตรวจสอบวันในสัปดาห์
        date_obj = datetime.strptime(day, "%d-%m-%Y")
        day_of_week = date_obj.strftime('%A')

        # ถ้าเป็นวันหยุดหรือวันเสาร์-อาทิตย์
        if day_of_week in ['Saturday', 'Sunday'] or date_obj.strftime("%Y-%m-%d") in holidays:
            # คำนวณเฉพาะ Total
            total = df_selected['To Home (kWh)'].sum()

            # เพิ่มค่า Total ในแถวที่ 28 (index 27)
            df_selected.loc[26, 'Time'] = 'Total'
            df_selected.loc[26, 'To Home (kWh)'] = total

            # เก็บข้อมูลใน summary_data สำหรับการสรุป
            summary_data.append([day, 0, 0, total])

        else:
            # คำนวณ Onpeak และ Offpeak สำหรับวันทำงาน
            df_selected['Onpeak'] = df_selected.apply(
                lambda row: row['To Home (kWh)'] if onpeak_start <= row['Time'] <= onpeak_end else 0, axis=1
            )
            df_selected['Offpeak'] = df_selected.apply(
                lambda row: row['To Home (kWh)'] if not onpeak_start <= row['Time'] <= onpeak_end else 0, axis=1
            )

            # คำนวณ Total
            onpeak_total = df_selected['Onpeak'].sum()
            offpeak_total = df_selected['Offpeak'].sum()
            total = df_selected['To Home (kWh)'].sum()

            # เพิ่มค่า Onpeak, Offpeak และ Total ในแถว 26, 27, และ 28
            df_selected.loc[24, 'Time'] = 'Onpeak'  # แถว 25 เป็น index 24 ใน pandas
            df_selected.loc[25, 'Time'] = 'Offpeak'
            df_selected.loc[26, 'Time'] = 'Total'
            df_selected.loc[24, 'To Home (kWh)'] = onpeak_total
            df_selected.loc[25, 'To Home (kWh)'] = offpeak_total
            df_selected.loc[26, 'To Home (kWh)'] = total

            # เก็บข้อมูลใน summary_data สำหรับการสรุป
            summary_data.append([day, onpeak_total, offpeak_total, 0])

        # เปลี่ยนชื่อคอลัมน์
        df_selected.rename(columns={'To Home (kWh)': f'Day {day}'}, inplace=True)

        # เก็บข้อมูลในลิสต์
        all_days_data.append(df_selected[['Time', f'Day {day}']])

# รวมข้อมูลทั้งหมดเข้าด้วยกัน
result_df = pd.concat(all_days_data, axis=1)

# ลบคอลัมน์ 'Time' ที่ซ้ำกัน
result_df = result_df.loc[:,~result_df.columns.duplicated()]

# บันทึกผลลัพธ์ลงในไฟล์ Excel
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    result_df.to_excel(writer, sheet_name='Sheet1', index=False)

# โหลดไฟล์ Excel เพื่อทำการปรับแต่งเพิ่มเติม
wb = load_workbook(output_file)
ws = wb['Sheet1']

# สร้างหัวตารางสรุปที่แถว 34
ws['C34'] = 'วันที่'
ws['D34'] = 'Onpeak'
ws['E34'] = 'Offpeak'
ws['F34'] = 'Holidays'
ws['G34'] = 'วันที่'
ws['H34'] = 'Onpeak'
ws['I34'] = 'Offpeak'
ws['J34'] = 'Holidays'

# จัดเรียง summary_data ตามลำดับวันที่
summary_data.sort(key=lambda x: datetime.strptime(x[0], "%d-%m-%Y"))

# กำหนดตำแหน่งแถวเริ่มต้นสำหรับข้อมูล
row_start = 35

# นำข้อมูลจาก summary_df ใส่ในตารางสรุปที่ Sheet1
for i, row in pd.DataFrame(summary_data).iterrows():
    if i < 15:
        col_offset = 0  # กลุ่มคอลัมน์แรกสำหรับวันที่ 1-15
        row_pos = row_start + i  # ข้อมูลอยู่ในแถวที่ 35-49
    else:
        col_offset = 4  # กลุ่มคอลัมน์ถัดไปสำหรับวันที่ 16-31
        row_pos = row_start + (i - 15)  # ข้อมูลอยู่ในแถวที่ 35-50 สำหรับกลุ่มวันที่ 16-31

    # กรอกข้อมูลลงในตาราง
    ws.cell(row=row_pos, column=3 + col_offset).value = row[0]  # วันที่
    ws.cell(row=row_pos, column=4 + col_offset).value = row[1]  # Onpeak
    ws.cell(row=row_pos, column=5 + col_offset).value = row[2]  # Offpeak
    ws.cell(row=row_pos, column=6 + col_offset).value = row[3]  # Holidays

    # ไฮไลต์สีถ้าวันนั้นเป็นวันหยุดหรือวันเสาร์-อาทิตย์
    date_obj = datetime.strptime(row[0], "%d-%m-%Y")
    day_of_week = date_obj.strftime('%A')

    if day_of_week in ['Saturday', 'Sunday']:
        fill = weekend_fill
    elif date_obj.strftime("%Y-%m-%d") in holidays:
        # ใช้สีเพิ่มเติมสำหรับวันหยุดที่ตรงกับวันธรรมดาใน MEA
        if selection == "MEA" and day_of_week not in ['Saturday', 'Sunday']:
            fill = mea_weekday_holiday_fill
        else:
            fill = holiday_fill
    else:
        fill = None
    
    if fill:
        for col in range(3 + col_offset, 7 + col_offset):  # ไฮไลต์สีสำหรับคอลัมน์วันที่, Onpeak, Offpeak, Holidays
            ws.cell(row=row_pos, column=col).fill = fill

# สรุปผลรวมที่ด้านล่างตาราง
ws['E54'] = 'Onpeak'
ws['E55'] = sum([row[1] for row in summary_data])

ws['F54'] = 'Offpeak'
ws['F55'] = sum([row[2] for row in summary_data])

ws['G54'] = 'Holiday'
ws['G55'] = sum([row[3] for row in summary_data])

ws['H54'] = 'Total'
ws['H55'] = sum([row[1] + row[2] + row[3] for row in summary_data])

# ปรับการจัดแนวให้อยู่ตรงกลาง
for row in ws.iter_rows(min_row=34, max_row=55, min_col=3, max_col=10):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# บันทึกไฟล์ Excel ที่มีการแก้ไขแล้ว
wb.save(output_file)

print(f'Report generated successfully: {output_file}')

